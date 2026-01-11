#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
엑셀 파일에서 2024년 근로소득 간이세액표 데이터를 읽어 SQL 생성
"""

import openpyxl
import sys
import os

def parse_excel_and_generate_sql(excel_path, output_sql_path):
    """엑셀 파일을 읽어서 SQL INSERT 문 생성"""
    
    print(f"[INFO] 엑셀 파일 읽는 중: {excel_path}")
    
    try:
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        print(f"[OK] 시트 이름: {ws.title}")
        print(f"[INFO] 총 행 수: {ws.max_row}")
        
        # SQL 생성 시작
        sql_lines = []
        sql_lines.append("-- 2024년 근로소득 간이세액표 정확한 데이터로 업데이트")
        sql_lines.append("-- 출처: 2024년 근로소득에 대한 간이세액표(제189조 관련)")
        sql_lines.append("-- 엑셀 파일에서 자동 생성")
        sql_lines.append("")
        sql_lines.append("BEGIN;")
        sql_lines.append("")
        sql_lines.append("-- 기존 데이터 삭제")
        sql_lines.append("DELETE FROM income_tax_brackets_2026;")
        sql_lines.append("")
        
        # 헤더 찾기 (이상/미만이 있는 행)
        header_row = None
        for row_idx in range(1, min(20, ws.max_row + 1)):
            cell_val = str(ws.cell(row_idx, 1).value or "")
            if "이상" in cell_val or "월급여액" in cell_val:
                header_row = row_idx + 1  # 다음 행부터 데이터
                break
        
        if not header_row:
            print("[WARN] 헤더를 찾을 수 없습니다")
            return False
        
        print(f"[INFO] 데이터 시작 행: {header_row}")
        
        insert_count = 0
        
        # 데이터 행 처리
        for row_idx in range(header_row, ws.max_row + 1):
            # 첫 번째 열 (이상 - 소득 하한)
            income_from_cell = ws.cell(row_idx, 1).value
            if not income_from_cell:
                continue
            
            # 두 번째 열 (미만 - 소득 상한)
            income_to_cell = ws.cell(row_idx, 2).value
            if not income_to_cell:
                continue
            
            try:
                # 천원 단위를 원 단위로 변환
                if isinstance(income_from_cell, str):
                    income_from = int(income_from_cell.replace(',', '')) * 1000
                else:
                    income_from = int(income_from_cell) * 1000
                
                if isinstance(income_to_cell, str):
                    income_to = int(income_to_cell.replace(',', '')) * 1000
                else:
                    income_to = int(income_to_cell) * 1000
                
                # 10,000천원 초과 데이터는 함수로 처리하므로 스킵
                if income_from >= 10000000:
                    break
                
                # 11개의 공제대상 가족 수에 대한 세액 (3~13열)
                for dep_count in range(1, 12):  # 1~11명
                    col_idx = dep_count + 2  # 3열부터 시작
                    tax_cell = ws.cell(row_idx, col_idx).value
                    
                    # 세액 처리
                    tax_amount = 0
                    if tax_cell and str(tax_cell).strip() not in ['', '-']:
                        if isinstance(tax_cell, str):
                            tax_amount = int(tax_cell.replace(',', ''))
                        else:
                            tax_amount = int(tax_cell)
                    
                    # INSERT 문 추가
                    sql_lines.append(
                        f"INSERT INTO income_tax_brackets_2026 "
                        f"(income_from, income_to, dependent_count, tax_amount) "
                        f"VALUES ({income_from}, {income_to}, {dep_count}, {tax_amount});"
                    )
                    insert_count += 1
                
                if insert_count % 110 == 0:  # 10개 구간마다 (11명 x 10 = 110)
                    print(f"  처리 중... {insert_count} rows")
            
            except Exception as e:
                print(f"[WARN] 행 {row_idx} 처리 오류: {e}")
                continue
        
        sql_lines.append("")
        sql_lines.append("COMMIT;")
        sql_lines.append("")
        sql_lines.append("-- [OK] 총 " + str(insert_count) + "개 레코드 삽입")
        sql_lines.append("")
        sql_lines.append("-- [INFO] 10,000천원 초과 구간은 calculate_income_tax_2026 함수에서 다음 수식으로 계산:")
        sql_lines.append("--   - 10,000~14,000천원: (10,000천원 세액) + (초과금액 * 0.98 * 0.35) + 25,000")
        sql_lines.append("--   - 14,000~28,000천원: (10,000천원 세액) + 1,397,000 + (14,000초과 * 0.98 * 0.38)")
        sql_lines.append("--   - 28,000~30,000천원: (10,000천원 세액) + 6,610,600 + (28,000초과 * 0.98 * 0.40)")
        sql_lines.append("--   - 30,000~45,000천원: (10,000천원 세액) + 7,394,600 + (30,000초과 * 0.40)")
        sql_lines.append("--   - 45,000~87,000천원: (10,000천원 세액) + 13,394,600 + (45,000초과 * 0.42)")
        sql_lines.append("--   - 87,000천원 초과: (10,000천원 세액) + 31,034,600 + (87,000초과 * 0.45)")
        
        # SQL 파일 저장
        with open(output_sql_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(sql_lines))
        
        print(f"\n[OK] SQL 파일 생성 완료: {output_sql_path}")
        print(f"[INFO] 총 {insert_count}개 레코드 INSERT 문 생성")
        
        return True
        
    except Exception as e:
        print(f"[ERROR] 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    # 엑셀 파일 경로
    excel_path = r'c:\Users\신치훈\Downloads\근로소득_간이세액표(조견표).xlsx'
    output_sql_path = 'database/fix_income_tax_2026_from_excel.sql'
    
    if not os.path.exists(excel_path):
        print(f"[ERROR] 엑셀 파일을 찾을 수 없습니다: {excel_path}")
        sys.exit(1)
    
    success = parse_excel_and_generate_sql(excel_path, output_sql_path)
    
    if success:
        print("\n[SUCCESS] SQL 생성 완료! 이제 Supabase에서 실행하세요.")
    else:
        print("\n[ERROR] SQL 생성 실패")
        sys.exit(1)
